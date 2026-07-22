"""
EduVision - Higher Education Dashboard
Module 3 : Education KPI Engineering

Reads data/university_cleaned.csv (Module 2 output) and engineers the
six KPIs called for in the project plan, saving the result as
data/university_final_dataset.xlsx - the Tableau-ready deliverable.

KPIs generated
--------------
1. Global Ranking Score        - world_rank normalized to a 0-100
                                  scale within each ranking year.
2. Research Impact Score       - blend of citation performance,
                                  research output and h-index.
3. Faculty-to-Student Ratio    - students per faculty member
                                  (existing metric, cleaned/capped).
4. International Student %     - share of the student body that is
                                  international (cleaned/capped 0-100).
5. Academic Reputation Score   - carried through from source data
                                  (already a 0-100 index).
6. Research Productivity Index - composite of publication volume,
                                  citation efficiency and h-index.

Author : Chandana P
"""

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"

INPUT_FILE = DATA_DIR / "university_cleaned.csv"
OUTPUT_FILE = DATA_DIR / "university_final_dataset.xlsx"


# ==========================================================
# HELPERS
# ==========================================================

def minmax_scale(series, lower_q=0.01, upper_q=0.99):
    """
    Scale a numeric series to 0-100. Clips to the 1st/99th percentile
    first so a handful of extreme outliers (e.g. a mega-university's
    publication count) don't compress everyone else into a sliver of
    the range.
    """

    lower = series.quantile(lower_q)
    upper = series.quantile(upper_q)

    clipped = series.clip(lower=lower, upper=upper)

    if upper == lower:
        return pd.Series(50.0, index=series.index)

    return ((clipped - lower) / (upper - lower) * 100).round(1)


# ==========================================================
# KPI FUNCTIONS
# ==========================================================

def add_global_ranking_score(df):
    """
    world_rank -> 0-100, normalized within each ranking year since
    the size of the ranked pool grows year over year (800 unis in
    2016 vs 2,191 in 2026).
    """

    max_rank_per_year = df.groupby("year")["world_rank"].transform("max")

    df["global_ranking_score"] = (
        100 * (1 - (df["world_rank"] - 1) / (max_rank_per_year - 1))
    ).round(1)

    return df


def add_research_impact_score(df):
    """
    Blend of the four research-quality signals available per row:
    citations_score and research_output_score are already 0-100;
    citations_per_faculty and h_index are min-max scaled onto the
    same range before averaging.
    """

    norm_citations_per_faculty = minmax_scale(df["citations_per_faculty"])
    norm_h_index = minmax_scale(df["h_index"])

    df["research_impact_score"] = (
        df[["citations_score", "research_output_score"]].mean(axis=1)
        * 0.5
        + norm_citations_per_faculty * 0.25
        + norm_h_index * 0.25
    ).round(1)

    return df


def add_faculty_student_ratio(df):
    """
    Students-per-faculty-member, taken from source data and capped at
    a realistic ceiling (100:1) so a handful of data-entry outliers
    don't distort dashboard axes.
    """

    df["faculty_to_student_ratio_kpi"] = df["faculty_to_student_ratio"].clip(
        lower=0, upper=100
    ).round(2)

    return df


def add_international_student_percentage(df):
    """
    international_student_ratio arrives as a percentage but a few
    rows exceed 100 due to source data errors; clip to a valid
    0-100 range.
    """

    df["international_student_percentage"] = df["international_student_ratio"].clip(
        lower=0, upper=100
    ).round(1)

    return df


def add_academic_reputation_score(df):
    """
    Already a 0-100 index in the cleaned data; carried through under
    a KPI-specific name for clarity in the final dataset.
    """

    df["academic_reputation_kpi"] = df["academic_reputation_score"].round(1)

    return df


def add_research_productivity_index(df):
    """
    Composite of publication volume (log-scaled, since raw counts
    span 19 to 400k+), citation efficiency and h-index - distinct
    from research_impact_score, which measures citation *quality*
    rather than research *output volume*.
    """

    norm_publications = minmax_scale(np.log1p(df["publications_count"]))
    norm_citations_per_faculty = minmax_scale(df["citations_per_faculty"])
    norm_h_index = minmax_scale(df["h_index"])

    df["research_productivity_index_kpi"] = (
        norm_publications * 0.5
        + norm_citations_per_faculty * 0.3
        + norm_h_index * 0.2
    ).round(1)

    return df


# ==========================================================
# EXCEL FORMATTING
# ==========================================================

def format_workbook(path):
    """
    Bold header row, professional font, frozen header, autofit-ish
    column widths. Pure formatting - all values were already written
    by pandas.to_excel, so there is nothing to recalculate.
    """

    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    body_font = Font(name="Arial", size=10)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 10), 40)

    wb.save(path)


# ==========================================================
# PIPELINE
# ==========================================================

def main():
    print("=" * 60)
    print("EduVision - Education KPI Engineering Pipeline")
    print("=" * 60)

    print(f"Reading {INPUT_FILE.name}")
    df = pd.read_csv(INPUT_FILE, low_memory=False)

    print("Calculating Global Ranking Score...")
    df = add_global_ranking_score(df)

    print("Calculating Research Impact Score...")
    df = add_research_impact_score(df)

    print("Calculating Faculty-to-Student Ratio...")
    df = add_faculty_student_ratio(df)

    print("Calculating International Student Percentage...")
    df = add_international_student_percentage(df)

    print("Calculating Academic Reputation Score...")
    df = add_academic_reputation_score(df)

    print("Calculating Research Productivity Index...")
    df = add_research_productivity_index(df)

    kpi_columns = [
        "global_ranking_score",
        "research_impact_score",
        "faculty_to_student_ratio_kpi",
        "international_student_percentage",
        "academic_reputation_kpi",
        "research_productivity_index_kpi",
    ]

    missing_in_kpis = df[kpi_columns].isna().sum().sum()
    assert missing_in_kpis == 0, f"{missing_in_kpis} missing KPI values!"

    print("Saving Tableau-ready dataset...")
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="university_final_dataset")

    print("Formatting workbook...")
    format_workbook(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("KPI ENGINEERING COMPLETED SUCCESSFULLY")
    print("=" * 60)
    print(f"Rows    : {df.shape[0]}")
    print(f"Columns : {df.shape[1]} (6 new KPI columns added)")
    print(f"Missing values in KPI columns : {missing_in_kpis}")
    print(f"\nSaved To:\n{OUTPUT_FILE}")

    return df


if __name__ == "__main__":
    main()